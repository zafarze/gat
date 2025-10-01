from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from weasyprint import HTML
from django.template.loader import render_to_string
from collections import defaultdict
from ..forms import MonitoringFilterForm
from ..models import StudentResult, SchoolClass, Subject, ClassSubject

def _get_monitoring_context(get_params, request_user):
    """
    Вспомогательная функция, которая готовит данные для страницы Мониторинга и экспорта.
    Принимает GET-параметры и возвращает готовый контекст.
    """
    form = MonitoringFilterForm(get_params or None, user=request_user)
    table_headers = []
    table_rows = []
    title_details = {} 

    if form.is_valid():
        results_qs = StudentResult.objects.select_related(
            'student__school_class__school', 'gat_test'
        ).all()
        
        # --- ФИЛЬТР ДЛЯ ДИРЕКТОРА ШКОЛЫ ---
        if not request_user.is_superuser and hasattr(request_user, 'profile') and request_user.profile.role == 'SCHOOL_DIRECTOR':
            results_qs = results_qs.filter(student__school_class__school=request_user.profile.school)
        # --- КОНЕЦ ФИЛЬТРА ---

        gat_tests = form.cleaned_data.get('gat_tests')
        academic_year = form.cleaned_data.get('academic_year')
        quarter = form.cleaned_data.get('quarter')
        schools = form.cleaned_data.get('schools')
        subjects_filter = form.cleaned_data.get('subjects')
        school_classes = form.cleaned_data.get('school_classes')

        if academic_year:
            results_qs = results_qs.filter(gat_test__quarter__year=academic_year)
        if quarter:
            results_qs = results_qs.filter(gat_test__quarter=quarter)
        if schools:
            results_qs = results_qs.filter(student__school_class__school__in=schools)
        if school_classes.exists():
            selected_class_ids = list(school_classes.values_list('id', flat=True))
            subclass_ids = list(SchoolClass.objects.filter(parent__in=selected_class_ids).values_list('id', flat=True))
            all_relevant_class_ids = set(selected_class_ids + subclass_ids)
            results_qs = results_qs.filter(student__school_class_id__in=all_relevant_class_ids)
        if subjects_filter:
            subject_keys = [str(s.id) for s in subjects_filter]
            results_qs = results_qs.filter(scores__has_any_keys=subject_keys)

        if schools:
            title_details['schools'] = ", ".join([s.name for s in schools])
        if school_classes:
            title_details['classes'] = ", ".join([c.name for c in school_classes])
        if quarter:
            title_details['period'] = str(quarter)
        elif academic_year:
            title_details['period'] = str(academic_year)

        is_total_mode = gat_tests and '1' in gat_tests and '2' in gat_tests
        if is_total_mode:
            title_details['test_type'] = "GAT Total"
        elif gat_tests and '1' in gat_tests:
            title_details['test_type'] = "GAT-1"
        elif gat_tests and '2' in gat_tests:
            title_details['test_type'] = "GAT-2"
        
        subject_map = {s.id: s for s in Subject.objects.all()}
        
        header_subjects = []
        if subjects_filter.exists():
            header_subjects = sorted(list(subjects_filter), key=lambda s: s.name)
        else:
            all_subject_ids_in_results = set()
            temp_qs = results_qs.filter(gat_test__test_number__in=[1, 2]) if is_total_mode else results_qs
            for r in temp_qs:
                all_subject_ids_in_results.update(int(sid) for sid in r.scores.keys())
            header_subjects = sorted([subject_map[sid] for sid in all_subject_ids_in_results if sid in subject_map], key=lambda s: s.name)

        q_counts = {}
        first_result = results_qs.first()
        if first_result:
            ref_classes = SchoolClass.objects.filter(
                id__in=results_qs.values_list('student__school_class_id', flat=True).distinct()
            ).select_related('parent')

            class_subjects_qs = ClassSubject.objects.filter(
                school_class__in=ref_classes, 
                subject__in=header_subjects
            )
            for cs in class_subjects_qs:
                q_counts[cs.subject_id] = cs.number_of_questions
            
            parent_class_ids = {c.parent_id for c in ref_classes if c.parent_id}
            if parent_class_ids:
                missing_subject_ids = [s.id for s in header_subjects if s.id not in q_counts]
                if missing_subject_ids:
                    parent_class_subjects_qs = ClassSubject.objects.filter(
                        school_class_id__in=parent_class_ids,
                        subject_id__in=missing_subject_ids
                    ).order_by('subject_id').distinct('subject_id')

                    for cs in parent_class_subjects_qs:
                        q_counts[cs.subject_id] = cs.number_of_questions

        table_headers = []
        for subj in header_subjects:
            table_headers.append({
                'subject': subj,
                'q_count': q_counts.get(subj.id, 0)
            })
        
        if is_total_mode:
            student_data = defaultdict(lambda: {"student": None, "scores": defaultdict(int), "tests": set()})
            for r in results_qs.filter(gat_test__test_number__in=[1, 2]):
                student_data[r.student_id]["student"] = r.student
                student_data[r.student_id]["tests"].add(r.gat_test.test_number)
                for sid, answers in r.scores.items():
                    student_data[r.student_id]["scores"][int(sid)] += sum(answers)
            
            for data in student_data.values():
                if len(data["tests"]) == 2:
                    total_score = sum(data["scores"].values())
                    table_rows.append({
                        'student': data["student"], 'is_total': True,
                        'scores_by_subject': data["scores"], 'total_score': total_score,
                    })
        else:
            if gat_tests:
                results_qs = results_qs.filter(gat_test__test_number__in=gat_tests)

            for result in results_qs.distinct():
                scores_by_subject = {int(sid): sum(ans) for sid, ans in result.scores.items()}
                table_rows.append({
                    'student': result.student, 'result_obj': result, 'is_total': False,
                    'scores_by_subject': scores_by_subject, 'total_score': sum(scores_by_subject.values()),
                })
        
        table_rows.sort(key=lambda x: x['total_score'], reverse=True)

    return {
        'form': form,
        'table_headers': table_headers,
        'table_rows': table_rows,
        'has_results': bool(get_params),
        'title_details': title_details,
    }

@login_required
def monitoring_view(request):
    """Отображает страницу мониторинга."""
    context = _get_monitoring_context(request.GET, request.user) # <-- Передаем request.user
    context['title'] = 'Мониторинг'
    return render(request, 'monitoring/monitoring.html', context)

@login_required
def export_monitoring_pdf(request):
    """Экспортирует результаты мониторинга в PDF."""
    context = _get_monitoring_context(request.GET, request.user) # <-- Передаем request.user
    context['title'] = 'Отчет по мониторингу'
    html_string = render_to_string('monitoring/monitoring_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="monitoring_report.pdf"'
    HTML(string=html_string).write_pdf(response)
    return response

@login_required
def export_monitoring_excel(request):
    """Экспортирует результаты мониторинга в Excel с двухуровневым заголовком."""
    context = _get_monitoring_context(request.GET, request.user) # <-- Передаем request.user
    table_headers = context['table_headers']
    table_rows = context['table_rows']
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="monitoring_report.xlsx"'
    
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Мониторинг'
    
    header1 = ["№", "ФИО Студента", "Класс", "Тест"]
    for header_data in table_headers:
        header1.append(header_data['subject'].abbreviation or header_data['subject'].name)
    header1.append("Общий балл")
    sheet.append(header1)

    header2 = ["", "", "", ""]
    for header_data in table_headers:
        header2.append(f"({header_data['q_count']})")
    header2.append("")
    sheet.append(header2)

    for col in range(1, 5):
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        sheet.cell(row=1, column=col).alignment = Alignment(vertical='center')

    total_score_col = len(header1)
    sheet.merge_cells(start_row=1, start_column=total_score_col, end_row=2, end_column=total_score_col)
    sheet.cell(row=1, column=total_score_col).alignment = Alignment(vertical='center')

    for i, row_data in enumerate(table_rows, 1):
        row = [
            i,
            str(row_data['student']),
            str(row_data['student'].school_class),
            "GAT Total" if row_data['is_total'] else row_data['result_obj'].gat_test.name,
        ]
        for header_data in table_headers:
            score = row_data['scores_by_subject'].get(header_data['subject'].id, "—")
            row.append(score)
        
        row.append(row_data['total_score'])
        sheet.append(row)

    for col_idx, column_cells in enumerate(sheet.columns, 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for cell in column_cells:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width
        
    workbook.save(response)
    return response
